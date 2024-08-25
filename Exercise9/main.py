import openpyxl

spreadsheet_file = openpyxl.load_workbook("employees.xlsx")
working_sheet = spreadsheet_file["Sheet1"]

employees = []

# Read values from file
for row in range (2, working_sheet.max_row + 1):
    current_employee = {}
    name = working_sheet.cell(row, 1).value
    years_of_experience = working_sheet.cell(row, 2).value
    job_title = working_sheet.cell(row, 3).value
    date_of_birth = working_sheet.cell(row, 4).value
    current_employee["Name"] = name
    current_employee["Years of Experience"] = years_of_experience
    current_employee["Job Title"] = job_title
    current_employee["Date of Birth"] = date_of_birth
    employees.append(current_employee)

sorted_employees = sorted(employees, key=lambda employee: employee["Years of Experience"], reverse=True)

# Write sorted values into file
i = 2
for employee in sorted_employees:
    working_sheet.cell(i, 1).value = employee["Name"]
    working_sheet.cell(i, 2).value = employee["Years of Experience"]
    working_sheet.cell(i, 3).value = employee["Job Title"]
    working_sheet.cell(i, 4).value = employee["Date of Birth"]
    i += 1

spreadsheet_file.save("employees_sorted.xlsx")
